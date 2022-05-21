import sqlite3

import openpyxl
from django.db import models


class ProductCategory(models.Model):
    name = models.CharField(max_length=100, unique=True)

    def __str__(self):
        return self.name


class Product(models.Model):
    name = models.CharField(max_length=100)
    description = models.TextField()
    price = models.DecimalField(max_digits=20, decimal_places=2, default=0)
    quantity = models.PositiveIntegerField(default=0)
    category = models.ForeignKey(ProductCategory, on_delete=models.CASCADE)

    def __str__(self):
        return f' Продукт: {self.name}| Категория: {self.category.name}'

    def safe_delete_product(self):
        self.quantity = 0
        self.save()


def export_to_sqlite():
    connect = sqlite3.connect('db.sqlite3')
    cursor = connect.cursor()
    cursor.execute("""CREATE TABLE IF NOT EXISTS products_product 
                    (id int, name text, description text, price bigint, category_id int )""")

    book = openpyxl.open('product.xlsx', read_only=False)
    sheet = book.active

    for row in range(2, sheet.max_row + 1):
        data = []
        for col in range(1, sheet.max_column + 1):
            value = sheet.cell(row, col).value
            data.append(value)
        cursor.execute("INSERT INTO products_product VALUES (?, ?, ?, ?, ?);", (data[0], data[1], data[2], data[3],
                                                                                data[4]))
    connect.commit()
    connect.close()
